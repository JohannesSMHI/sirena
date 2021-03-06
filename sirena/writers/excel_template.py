# Copyright (c) 2020 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).
"""
Created on 2021-03-29 13:20
@author: johannes
"""
import os
import shutil
import openpyxl
from openpyxl import load_workbook
from sirena.writers.writer import WriterBase


class ExcelTemplateWriter(WriterBase):
    """
    """
    def __init__(self, template_path=None, export_path=None, cell_mapping=None, **kwargs):
        super(ExcelTemplateWriter, self).__init__()
        self.workbook_template_path = template_path
        self.export_path = export_path
        self.cell_mapping = cell_mapping
        self.template_sheetname = 'mwreg'
        self.workbook = None
        for name, item in kwargs.items():
            setattr(self, name, item)

        self.copy_new_woorkbook()

    def write(self, data):
        ws = self.workbook[self.template_sheetname]

        row_number = self.start_row_index or 0
        for statn, item in data.items():
            for attr, value in item.items():
                cell_id = self.cell_mapping.get(attr) % row_number
                ws[cell_id] = value
            row_number += 1
            if row_number == 50:
                row_number = 60

        img = openpyxl.drawing.image.Image(r'C:\Utveckling\sirena\sirena\etc\icons\smhi_sjov.png')
        img.anchor = 'B2'
        ws.add_image(img)
        self.workbook.template = False
        self.workbook.save(self.export_path)

    def copy_new_woorkbook(self):
        shutil.copyfile(self.workbook_template_path, self.export_path)
        self.workbook = load_workbook(self.export_path)

    def get_new_sheet(self, new_sheet_name):
        ws = self.workbook.copy_worksheet(self.workbook[self.template_sheetname])
        ws.title = new_sheet_name
        return ws


if __name__ == '__main__':
    wd = os.path.dirname(os.path.realpath(__file__))
    tmp_path = os.path.join(wd, 'etc', 'templates', 'mwreg.xlsx')
    # export_path = os.path.join(wd, 'export')

    handler = ExcelTemplateWriter(
        tmp_path=tmp_path,
        # export_path=export_path,
    )
